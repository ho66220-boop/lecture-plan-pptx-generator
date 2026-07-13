import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from config.defaults import REPORT_COLUMNS
except ModuleNotFoundError:
    from ..config.defaults import REPORT_COLUMNS


NORMALIZED_COLUMNS = [
    "source_file",
    "source_sheet",
    "lecture_id",
    "강좌 유형",
    "학년",
    "과목",
    "강사명",
    "강의형태",
    "구분",
    "시즌",
    "강좌명",
    "메인 제목",
    "서브 슬로건",
    "수업 요일 / 시간",
    "QC/RC 또는 클리닉 시간",
    "수강 대상",
    "휴강일 / 사유 / 수업 불가 일정",
    "보강 방법",
    "교재 정보",
    "강의 개요",
    "강의특징",
    "관리 프로그램",
    "수강 후기",
    "연간 커리큘럼",
    "비고 / 요청사항",
    "computed_opening_date",
    "opening_date_display",
    "computed_total_sessions",
    "computed_period",
    "computed_fee",
    "fee_per_session",
    "fee_display",
    "flags",
    "진도표",
]


def _save_workbook_with_fallback(wb, path, label):
    """산출 엑셀 저장. 파일이 엑셀에서 열려 있어 잠기면(PermissionError 등) 타임스탬프
    대체명으로 1회 재시도해 배치 결과를 최대한 파일로 남긴다(특히 validation_report는
    실패 원인을 알려줄 최후의 보루라 반드시 저장 시도). 그마저 실패하면 스택트레이스
    대신 원인·조치를 담은 RuntimeError로 명확히 실패한다."""
    try:
        wb.save(path)
        return path
    except Exception:
        alt = path.with_name(f"{path.stem}_재시도{datetime.now().strftime('%H%M%S')}{path.suffix}")
        try:
            wb.save(alt)
            return alt
        except Exception as exc:
            raise RuntimeError(
                f"{label} 파일을 저장할 수 없습니다: {path} ({type(exc).__name__}). "
                "산출 파일이 엑셀에서 열려 있는지 확인하고, 닫은 뒤 다시 실행해 주세요."
            ) from exc


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_len + 2, 12), 55)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_normalized_data(lectures, output_dir):
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)

    json_path = output / "normalized_data.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(lectures, f, ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "normalized_data"
    ws.append(NORMALIZED_COLUMNS)
    for lecture in lectures:
        fields = lecture.get("fields", {})
        row = []
        for column in NORMALIZED_COLUMNS:
            if column == "flags":
                row.append(", ".join(lecture.get("flags", [])))
            elif column == "진도표":
                row.append(json.dumps(lecture.get("progress", []), ensure_ascii=False))
            elif column in fields:
                row.append(fields.get(column, ""))
            else:
                row.append(lecture.get(column, ""))
        ws.append(row)
    style_sheet(ws)
    xlsx_path = _save_workbook_with_fallback(
        wb, output / "normalized_data.xlsx", "정규화 데이터"
    )
    return xlsx_path, json_path


def export_validation_report(reports, output_dir):
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "validation_report"
    ws.append(REPORT_COLUMNS)
    for item in reports:
        ws.append([item.get(column, "") for column in REPORT_COLUMNS])
    style_sheet(ws)
    path = _save_workbook_with_fallback(
        wb, output / "validation_report.xlsx", "검증 리포트"
    )
    return path
